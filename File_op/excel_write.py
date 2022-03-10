"""
通过pandas库将list或numpy写入excel
"""
import numpy as np
import pandas as pd
import os

from openpyxl import load_workbook


def write_excel(arrayOrList, savePath, saveName="name.xlsx"):
    """
    :param arrayOrList: 要保存的numpy数组或list列表，一维或二维
    :param savePath: 保存路径
    :param saveName: excel名命
    :return:

    example1：
        # 二维list
        company_name_list = [['腾讯', '北京'], ['阿里巴巴', '杭州'], ['字节跳动', '北京']]
        array = np.array(company_name_list)
        # list转dataframe
        df = pd.DataFrame(array, columns=['company_name', 'local'])
        # 保存到本地excel
        df.to_excel("company_name_li1.xlsx", index=False)
    example2（more correct）:
        company_name_list = [['腾讯', '北京'], ['阿里巴巴', '杭州'], ['字节跳动', '北京']]
        array = np.array(company_name_list)
        data = pd.DataFrame(array, columns=['company_name', 'local'])
        writer = pd.ExcelWriter(os.path.join(r"D:\MyProjects\PycharmProjects\video", "name.xlsx"))
        data.to_excel(writer, "page_1", float_format='%d', index=False)
        writer.save()
        writer.close()
    """

    # 转换数据格式，可以通过colums属性设置excel第一行的行头
    data = pd.DataFrame(arrayOrList)
    writer = pd.ExcelWriter(os.path.join(savePath, saveName))
    # 可以通过header= index= 属性设置是否需要行号和列号（True or False）
    # float_format = ’%d’ 存入的是整数格式， float_format =’%.5f’ 存入的是小数格式，小数点后可有5位
    data.to_excel(writer, sheet_name="Sheet1", float_format='%d', index=False)
    writer.save()
    writer.close()


def write_excel_add(arrayOrList_add, savePath, saveName="name.xlsx"):
    """
        :param arrayOrList_add: 要保存的numpy数组或list列表，一维或二维
        :param savePath: 保存路径
        :param saveName: excel名命
        :return:

    example:
        result2 = [('a', '2'), ('b', '2'), ('c', '4')]  # 需要新写入的数据
        df = pd.DataFrame(result2)  # 列表数据转为数据框
        filename = os.path.join(r"D:\MyProjects\PycharmProjects\video", "name.xlsx")
        df1 = pd.DataFrame(pd.read_excel(filename, sheet_name='page_1', engine='openpyxl'))  # 读取原数据文件和表
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        book = load_workbook(filename)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_rows = df1.shape[0]  # 获取原数据的行数
        df.to_excel(writer, sheet_name='page_1', startrow=df_rows + 2, index=False, header=False)  # 将数据写入excel中的aa表,从第一个空行开始写
        writer.save()  # 保存
    """
    # 数据转换为数据框
    data_add = pd.DataFrame(arrayOrList_add)
    filename = os.path.join(savePath,saveName)
    # 读取原数据文件和表
    data_read = pd.DataFrame(pd.read_excel(filename, sheet_name='Sheet1', engine='openpyxl'))
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    book = load_workbook(filename)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # 获取原数据的行数
    df_rows = data_read.shape[0]
    data_add.to_excel(writer, sheet_name='page_1', startrow=df_rows + 2, index=False,
                      header=False)  # 将数据写入excel中的aa表,从第一个空行开始写
    # 保存
    writer.save()
